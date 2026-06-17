#!/usr/bin/env python3
"""Generate the Ice Business Ledger XLSX template (import to Google Sheets)."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

OUTPUT = Path(__file__).resolve().parent.parent / "templates" / "ice-business-ledger.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROUTES = [f"Route_{i:02d}" for i in range(1, 7)]
STORES = [f"Store_{i:02d}" for i in range(1, 11)]
EXPENSE_GROUPS = {
    "Production": ["Water", "Power", "Bags", "Machine repair"],
    "Distribution": ["Fuel", "Vehicle repair", "Driver pay"],
    "Overhead": ["Rent", "Permits", "Insurance", "Other"],
}
EXPENSE_CATS = [c for cats in EXPENSE_GROUPS.values() for c in cats]

# Workbook-level named ranges (Google Sheets imports these more reliably than
# per-cell cross-sheet formulas in data validation).
NAMED_RANGES = {
    "Routes": "Settings!$A$2:$A$100",
    "Stores": "Settings!$C$2:$C$500",
    "ExpenseCategories": "Settings!$E$2:$E$100",
    "SaleTypes": "Settings!$G$2:$G$3",
    "PaymentMethods": "Settings!$H$2:$H$3",
}


def style_header_row(ws, row, cols, fill=HEADER_FILL):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_named_ranges(wb):
    for name, refers_to in NAMED_RANGES.items():
        wb.defined_names[name] = DefinedName(name=name, attr_text=refers_to)


def add_list_validation(ws, cell_range, source):
    """Apply one list validation to an entire column range.

    `source` is a workbook named range (e.g. "Routes") or a Settings range
    formula (e.g. "=Settings!$A$2:$A$100"). Named ranges survive Google
  Sheets import better than thousands of per-cell cross-sheet rules.
    """
    formula = source if source.startswith("=") else f"={source}"
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,  # OOXML quirk: False shows the dropdown arrow
    )
    dv.error = "Pick from the list"
    dv.errorTitle = "Invalid value"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def build_settings(wb):
    ws = wb.active
    ws.title = "Settings"

    ws["A1"] = "Routes / เส้นทาง"
    ws["C1"] = "Stores / ร้านค้า"
    ws["E1"] = "Expense categories / หมวดค่าใช้จ่าย"
    ws["G1"] = "Sale type / ประเภทการขาย"
    ws["H1"] = "Payment method / ช่องทางจ่าย"
    style_header_row(ws, 1, 8)

    for i, route in enumerate(ROUTES, 2):
        ws.cell(row=i, column=1, value=route)
    for i, store in enumerate(STORES, 2):
        ws.cell(row=i, column=3, value=store)
    row = 2
    for group, cats in EXPENSE_GROUPS.items():
        for cat in cats:
            ws.cell(row=row, column=5, value=cat)
            ws.cell(row=row, column=6, value=group)
            row += 1
    ws["G2"], ws["G3"] = "Cash", "Credit"
    ws["H2"], ws["H3"] = "Cash", "Bank"

    set_widths(ws, [14, 4, 22, 4, 18, 14, 12, 14])


def build_route_sales(wb):
    ws = wb.create_sheet("Route_Sales")
    headers = [
        "Date / วันที่",
        "Route / เส้นทาง",
        "Store / ร้าน",
        "Bags / ถุง",
        "Sale type / สดหรือค้างจ่าย",
        "Amount (฿) / จำนวนเงิน",
        "Notes / หมายเหตุ",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    sample = [
        ("2026-06-16", "Route_01", "Store_01", 10, "Cash", 300, ""),
        ("2026-06-16", "Route_01", "Store_02", 8, "Credit", 240, ""),
    ]
    for r, row in enumerate(sample, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    add_list_validation(ws, "B2:B501", "Routes")
    add_list_validation(ws, "C2:C501", "Stores")
    add_list_validation(ws, "E2:E501", "SaleTypes")

    set_widths(ws, [12, 14, 20, 8, 22, 14, 24])


def build_store_payments(wb):
    ws = wb.create_sheet("Store_Payments")
    headers = [
        "Date / วันที่",
        "Store / ร้าน",
        "Amount paid (฿) / จำนวนที่จ่าย",
        "Payment method / ช่องทางจ่าย",
        "Notes / หมายเหตุ",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    ws.append(["2026-06-20", "Store_02", 240, "Cash", "Paid in full"])

    add_list_validation(ws, "B2:B501", "Stores")
    add_list_validation(ws, "D2:D501", "PaymentMethods")

    set_widths(ws, [12, 22, 18, 18, 24])


def build_store_balances(wb):
    ws = wb.create_sheet("Store_Balances")
    ws["A1"] = "Read-only / อ่านอย่างเดียว — who owes what / ลูกค้าค้างจ่าย"
    ws.merge_cells("A1:E1")
    ws["A1"].font = Font(bold=True, size=11)
    ws["A1"].fill = SUBHEADER_FILL

    headers = [
        "Store / ร้าน",
        "Opening debt (฿) / ยอดยกมา",
        "Credit sales (฿) / ขายเชื่อ",
        "Payments (฿) / รับชำระ",
        "Still owed (฿) / คงค้าง",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, len(headers))

    for i, store in enumerate(STORES, 3):
        r = i
        ws.cell(row=r, column=1, value=store)
        ws.cell(
            row=r,
            column=2,
            value=f'=IFERROR(SUMIF(Opening_Balances!$A:$A,A{r},Opening_Balances!$B:$B),0)',
        )
        ws.cell(
            row=r,
            column=3,
            value=f'=SUMIFS(Route_Sales!$F:$F,Route_Sales!$C:$C,A{r},Route_Sales!$E:$E,"Credit")',
        )
        ws.cell(row=r, column=4, value=f"=SUMIF(Store_Payments!$B:$B,A{r},Store_Payments!$C:$C)")
        ws.cell(row=r, column=5, value=f"=B{r}+C{r}-D{r}")

    set_widths(ws, [22, 16, 16, 16, 16])


def build_expenses(wb):
    ws = wb.create_sheet("Expenses")
    headers = [
        "Date / วันที่",
        "Category / หมวด",
        "Amount (฿) / จำนวนเงิน",
        "Payment / ช่องทางจ่าย",
        "Vendor or notes / รายละเอียด",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    ws.append(["2026-06-16", "Fuel", 800, "Cash", "Route 01"])

    add_list_validation(ws, "B2:B501", "ExpenseCategories")
    add_list_validation(ws, "D2:D501", "PaymentMethods")

    set_widths(ws, [12, 18, 14, 14, 28])


def build_inventory(wb):
    ws = wb.create_sheet("Inventory")
    headers = [
        "Week ending / สิ้นสัปดาห์",
        "Bags produced / ผลิต",
        "Bags sold / ขาย",
        "Variance / ต่าง",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    ws["A2"] = "2026-06-15"
    ws["B2"] = 500
    # Bags sold in the 7-day week ending on column A (Sun–Sat if A is Sunday)
    ws["C2"] = '=SUMIFS(Route_Sales!$D:$D,Route_Sales!$A:$A,">="&A2-6,Route_Sales!$A:$A,"<="&A2)'
    ws["D2"] = "=B2-C2"

    set_widths(ws, [16, 14, 14, 12])


def build_opening_balances(wb):
    ws = wb.create_sheet("Opening_Balances")
    ws["A1"] = "Go-live snapshot / ยอดยกมาวันเริ่มใช้"
    ws.merge_cells("A1:B1")
    ws["A1"].font = Font(bold=True)
    ws["A1"].fill = SUBHEADER_FILL

    ws["A3"], ws["B3"] = "Cash on hand / เงินสด", 0
    ws["A4"], ws["B4"] = "Bank balance / เงินในบัญชี", 0
    ws["A6"] = "Per-store opening debt / ยอดค้างจ่ายยกมา (รายร้าน)"
    ws["A6"].font = Font(bold=True)

    headers = ["Store / ร้าน", "Opening debt (฿) / ยอดค้าง"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=7, column=col, value=h)
    style_header_row(ws, 7, 2, fill=SUBHEADER_FILL)

    for i, store in enumerate(STORES, 8):
        ws.cell(row=i, column=1, value=store)
        ws.cell(row=i, column=2, value=0)

    add_list_validation(ws, "A8:A107", "Stores")

    set_widths(ws, [28, 18])


def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws["A1"] = "Monthly summary / สรุปรายเดือน"
    ws["A1"].font = Font(bold=True, size=12)

    ws["A3"] = "Report month (YYYY-MM) / เดือน:"
    ws["B3"] = "2026-06"

    labels = [
        (
            "Route cash income / รายได้เงินสด",
            '=SUMIFS(Route_Sales!$F:$F,Route_Sales!$E:$E,"Cash",Route_Sales!$A:$A,">="&DATE(LEFT($B$3,4),RIGHT($B$3,2),1),Route_Sales!$A:$A,"<"&EDATE(DATE(LEFT($B$3,4),RIGHT($B$3,2),1),1))',
        ),
        (
            "Route credit income / รายได้ขายเชื่อ",
            '=SUMIFS(Route_Sales!$F:$F,Route_Sales!$E:$E,"Credit",Route_Sales!$A:$A,">="&DATE(LEFT($B$3,4),RIGHT($B$3,2),1),Route_Sales!$A:$A,"<"&EDATE(DATE(LEFT($B$3,4),RIGHT($B$3,2),1),1))',
        ),
        ("Total route income / รายได้รวม", "=B5+B6"),
        (
            "Total expenses / ค่าใช้จ่ายรวม",
            '=SUMIFS(Expenses!$C:$C,Expenses!$A:$A,">="&DATE(LEFT($B$3,4),RIGHT($B$3,2),1),Expenses!$A:$A,"<"&EDATE(DATE(LEFT($B$3,4),RIGHT($B$3,2),1),1))',
        ),
        ("Net (income - expenses) / กำไรสุทธิ", "=B7-B8"),
    ]
    row = 5
    for label, formula in labels:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=formula)
        row += 1

    ws["A11"] = "Top stores still owed / ร้านค้างจ่ายสูงสุด"
    ws["A11"].font = Font(bold=True)
    ws["A12"] = "See Store_Balances tab / ดูแท็บ Store_Balances"

    ws["A14"] = "Expense breakdown / ค่าใช้จ่ายตามหมวด"
    ws["A14"].font = Font(bold=True)
    r = 15
    for cat in EXPENSE_CATS:
        ws.cell(row=r, column=1, value=cat)
        ws.cell(
            row=r,
            column=2,
            value=f'=SUMIFS(Expenses!$C:$C,Expenses!$B:$B,"{cat}",Expenses!$A:$A,">="&DATE(LEFT($B$3,4),RIGHT($B$3,2),1),Expenses!$A:$A,"<"&EDATE(DATE(LEFT($B$3,4),RIGHT($B$3,2),1),1))',
        )
        r += 1

    set_widths(ws, [36, 18])


def build_cash_flow(wb):
    ws = wb.create_sheet("Cash_Flow")
    ws["A1"] = "Cash flow / กระแสเงินสด"
    ws["A1"].font = Font(bold=True, size=12)

    ws["A3"] = "Report month (YYYY-MM):"
    ws["B3"] = "2026-06"

    items = [
        ("Opening cash / เงินสดยกมา", "=Opening_Balances!B3"),
        ("Opening bank / เงินธนาคารยกมา", "=Opening_Balances!B4"),
        (
            "+ Route cash collected / + เงินสดจากเส้นทาง",
            '=SUMIFS(Route_Sales!$F:$F,Route_Sales!$E:$E,"Cash",Route_Sales!$A:$A,">="&DATE(LEFT($B$3,4),RIGHT($B$3,2),1),Route_Sales!$A:$A,"<"&EDATE(DATE(LEFT($B$3,4),RIGHT($B$3,2),1),1))',
        ),
        (
            "+ Store payments (cash) / + รับชำระเงินสด",
            '=SUMIFS(Store_Payments!$C:$C,Store_Payments!$D:$D,"Cash",Store_Payments!$A:$A,">="&DATE(LEFT($B$3,4),RIGHT($B$3,2),1),Store_Payments!$A:$A,"<"&EDATE(DATE(LEFT($B$3,4),RIGHT($B$3,2),1),1))',
        ),
        (
            "+ Store payments (bank) / + รับชำระโอน",
            '=SUMIFS(Store_Payments!$C:$C,Store_Payments!$D:$D,"Bank",Store_Payments!$A:$A,">="&DATE(LEFT($B$3,4),RIGHT($B$3,2),1),Store_Payments!$A:$A,"<"&EDATE(DATE(LEFT($B$3,4),RIGHT($B$3,2),1),1))',
        ),
        (
            "- Cash expenses / - จ่ายเงินสด",
            '=SUMIFS(Expenses!$C:$C,Expenses!$D:$D,"Cash",Expenses!$A:$A,">="&DATE(LEFT($B$3,4),RIGHT($B$3,2),1),Expenses!$A:$A,"<"&EDATE(DATE(LEFT($B$3,4),RIGHT($B$3,2),1),1))',
        ),
        (
            "- Bank expenses / - จ่ายโอน/บัญชี",
            '=SUMIFS(Expenses!$C:$C,Expenses!$D:$D,"Bank",Expenses!$A:$A,">="&DATE(LEFT($B$3,4),RIGHT($B$3,2),1),Expenses!$A:$A,"<"&EDATE(DATE(LEFT($B$3,4),RIGHT($B$3,2),1),1))',
        ),
    ]
    row = 5
    for label, formula in items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=formula)
        row += 1

    ws.cell(row=row, column=1, value="Expected cash on hand / เงินสดคาดว่า")
    ws.cell(row=row, column=2, value="=B5+B7-B10")
    row += 1
    ws.cell(row=row, column=1, value="Expected bank balance / เงินธนาคารคาดว่า")
    ws.cell(row=row, column=2, value="=B6+B8+B9-B11")

    set_widths(ws, [40, 18])


def build_readme_setup(wb):
    """In-sheet instructions if Google Sheets drops validation on import."""
    ws = wb.create_sheet("README_Setup", 0)
    ws["A1"] = "Google Sheets dropdown setup / ตั้งค่า dropdown"
    ws["A1"].font = Font(bold=True, size=12)

    lines = [
        "",
        "After importing this file into Google Sheets, click a Route or Store cell.",
        "If you do NOT see a dropdown arrow, re-apply validation once per column:",
        "",
        "Route_Sales",
        "  Column B (Route):     Data → Data validation → Dropdown from a range → Routes",
        "  Column C (Store):     … → Stores",
        "  Column E (Sale type): … → SaleTypes",
        "",
        "Store_Payments",
        "  Column B (Store):     … → Stores",
        "  Column D (Payment):   … → PaymentMethods",
        "",
        "Expenses",
        "  Column B (Category):  … → ExpenseCategories",
        "  Column D (Payment):   … → PaymentMethods",
        "",
        "Opening_Balances",
        "  Column A (Store):     … → Stores   (rows 8–107)",
        "",
        "Named ranges (Data → Named ranges) point at Settings lists.",
        "Add new route/store names on Settings — dropdowns update automatically.",
        "",
        "Or run Extensions → Apps Script → paste scripts/google-sheets-validation.gs",
        "then Run → applyLedgerValidations (one-time fix).",
    ]
    for i, line in enumerate(lines, 2):
        ws.cell(row=i, column=1, value=line)

    ws.column_dimensions["A"].width = 72


def main():
    wb = Workbook()
    build_settings(wb)
    add_named_ranges(wb)
    build_route_sales(wb)
    build_store_payments(wb)
    build_store_balances(wb)
    build_expenses(wb)
    build_inventory(wb)
    build_opening_balances(wb)
    build_dashboard(wb)
    build_cash_flow(wb)
    build_readme_setup(wb)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
